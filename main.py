# -*- coding: utf-8 -*-
import csv
import os
import sys
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import Font, colors, Border, Side, PatternFill, Alignment
import time
import numpy as np
def get_file_path():
    # 获取脚本路径
    path = sys.path[0]
    # 判断为脚本文件还是py2exe编译后的文件，如果是脚本文件，则返回的是脚本的目录，如果是py2exe编译后的文件，则返回的是编译后的文件路径
    if os.path.isdir(path):
        return path
    elif os.path.isfile(path):
        return os.path.dirname(path)

RUSHOUSHIJIAN = 7   #入手条件：连续n天
LIANXU = 2          #连续上涨天数
HUANSHOULV = -1     #
CHUSHOUSHIJIAN = 3  #出手条件

class ReadAllCSV():
    #获取文档路径
    def getfile(self):
        path = os.path.join(get_file_path(), "resource\Data")
        files = os.listdir(path)
        for filenames in files:
            file_name = os.path.join(get_file_path(), "resource\Data\{}".format(filenames))
            yield file_name
        yield None

    #获取单份文档详细内容
    def getSigleCSV(self, name):
        file_name = os.path.join(get_file_path(), "resource\Data\data{}.csv".format(name))
        if os.path.exists(file_name):
            data = self.getSinglePageContent(file_name)
            data_in, data_out, round_lst = self.getTargeInAndOut(data)
            self.WriteSigleSummary(data_in, data_out, round_lst)
        else:
            print("文档不存在")
    #获取总涨幅文档
    def getTotalIncreaseCSV(self):
        totalIncrease_lst = []
        f = self.getfile()
        try:
            while True:
                file_name = f.__next__()
                print("filename:", file_name)
                if file_name:
                    data = self.getSinglePageContent(file_name)
                    data_in, data_out, round_lst = self.getTargeInAndOut(data)
                    # print("data_in", data_in)
                    if len(data_in) == 0:
                        one_stock_lst = [file_name.split("\\")[-1], "", -999]
                    else:
                        one_stock_lst = self.getTotalIncrease(data_in, data_out)
                    totalIncrease_lst.insert(0, one_stock_lst)
        except StopIteration as s:
            print(s.__repr__())
        
        self.writeTotalSummary(totalIncrease_lst)

    #获取每份文档的内容
    def getSinglePageContent(self, filename):
        if filename:
                with open(file=filename, encoding='ANSI') as fd:
                    df = pd.read_csv(fd, sep=',', encoding='utf-8', usecols=['日期', '股票代码', '名称', '开盘价', '换手率'])
                    df['日期'] = pd.to_datetime(df['日期'], format="%Y-%m-%d")
                    df.sort_values(by='日期', inplace=True)
                    # print(df)
                    return df
        else:
            return None
    #获取买入、卖出的数据
    def getTargeInAndOut(self, data):
        target_in = []              #锁定买入的几天的数组
        ishold = False              #是否持股
        target_out = []             #锁定卖出的几天的数组
        in_comparision_lst = []     #买入时用来比对的数组
        rise_lst = []               #连续上升的几天的数组
        out_comparision_lst = []    #卖出时用来比对的数组
        round_lst = []               #买卖一回合的数组，不包括卖出那天
        for indexs in data.index: 
            data_one_day = data.loc[indexs].values
            # print("data:", data_one_day)
            
            #个别无数据的跳过
            if  data_one_day[3] == "None"  or data_one_day[4] == "None":
                continue
            # print("1:", type(data_one_day[3]), "2:", type(data_one_day[4]))
            isrise = False
            in_comparision_lst.insert(0, data_one_day)
            out_comparision_lst.insert(0, data_one_day)

            if not ishold:
                #objective:用来比较的对象,RUSHOUSHIJIAN + 1是为了表示拿第8天与前7（RUSHOUSHIJIAN）天做比较
                if len(in_comparision_lst) == RUSHOUSHIJIAN + 1:
                    total = 0
                    for i in range(1, len(in_comparision_lst)):
                        total += float(in_comparision_lst[i][HUANSHOULV])
                    average = total/RUSHOUSHIJIAN
                    
                    #判断换手率大于7天平均值
                    if float(data_one_day[HUANSHOULV]) > average and average != 0:
                        rise_lst.insert(0, data_one_day)
                        isrise = True   
                    #买入 确保买入的时间是连续4天上升的第二天
                    if len(rise_lst) == LIANXU + 1:
                        target_in.insert(0, data_one_day)
                        ishold = True
                        rate_lst_one_round = [data_one_day]
                    if not isrise or ishold:
                        rise_lst.clear()
            else:
                if len(out_comparision_lst) == CHUSHOUSHIJIAN+1:
                    out_totle = 0
                    #从1开始 因为0是 用来比较的目标对象
                    for j in range(1, len(out_comparision_lst)):
                        out_totle += float(out_comparision_lst[j][HUANSHOULV])
                    out_average = out_totle/CHUSHOUSHIJIAN

                    if float(data_one_day[HUANSHOULV]) < out_average and out_average != 0:
                        target_out.insert(0, data_one_day)
                        ishold = False
                        round_lst.insert(0, rate_lst_one_round)
                if ishold:
                    rate_lst_one_round.insert(0, data_one_day)
            if len(in_comparision_lst) == RUSHOUSHIJIAN + 1:
                del in_comparision_lst[-1]  #删除最后一个元素
            if len(out_comparision_lst) == CHUSHOUSHIJIAN+1:
                del out_comparision_lst[-1]
        return target_in, target_out, round_lst
    #获取总涨幅
    def getTotalIncrease(self, data_in, data_out):
        gain_totl = 0
        for i in range(0, len(data_in)):
            if len(data_in) == len(data_out):
                # print(data_out[i])
                gain = np.float64(data_out[i][3]).item() - np.float64(data_in[i][3]).item()
                if np.float64(data_in[i][3]).item() == 0:
                    gain_totl = 0
                else:
                    print("data_in[i][3]:", np.float64(data_in[i][3]).item())
                    gain = gain / np.float64(data_in[i][3]).item()
                    gain_totl += gain
            elif len(data_in) > len(data_out):
                gain = np.float64(data_out[i-1][3]).item() - np.float64(data_in[i][3]).item()
                gain = gain / np.float64(data_in[i][3]).item()
                gain_totl += gain
        
        # print("代码:", data_in[0][1], "名称:", data_in[0][2], "总涨幅:", gain_totl)
        one_stock_lst = [data_in[0][1], data_in[0][2], gain_totl]

        return one_stock_lst
    #写所有文档的总结
    def writeTotalSummary(self, total_increase):
        total_increase.sort(key=lambda x:x[2],reverse=True)
        col_titles = ["A", "B", "C"]
        col_widthes = [8, 12, 12]
        dst_wb = openpyxl.Workbook()
        dst_sheet = dst_wb.active
        dst_sheet.title = u"股票统计"
        dst_sheet['A1'] = u"股票代码"
        dst_sheet['B1'] = u"名称"
        dst_sheet['C1'] = u"总涨幅"

        for i in range(0, len(total_increase)):
            dst_sheet[u"A{}".format(2+i)] = total_increase[i][0].split("'")[-1]
            dst_sheet[u"B{}".format(2+i)] = total_increase[i][1]
            dst_sheet[u"C{}".format(2+i)] = total_increase[i][2]

        font = Font(u"宋体", size=12)
        for col in range(0, len(col_titles)):
            dst_sheet.column_dimensions["{}".format(col_titles[col])].width = col_widthes[col]
            for row in range(1, dst_sheet.max_row + 1):
                dst_sheet[u"{}{}".format(col_titles[col], row)].font = font
                dst_sheet[u"{}{}".format(col_titles[col], row)].alignment = Alignment(horizontal='center', vertical='center')

        filename = '002.xlsx'
        dst_wb.save(filename)
    #写入单份文档的总结
    def WriteSigleSummary(self, data_in, data_out, round_lst):
        col_titles = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
        col_widthes = [8, 12, 12, 12, 12, 12, 12, 12,12, 12, 12]

        dst_wb = openpyxl.Workbook()
        dst_sheet = dst_wb.active
        dst_sheet.title = u"股票统计"
        dst_sheet['A2'] = u"股票代码"
        dst_sheet['B2'] = u"名称"
        dst_sheet['C1'] = u"买入"
        dst_sheet.merge_cells("C1:E1")
        dst_sheet['C2'] = u"日期"
        dst_sheet['D2'] = u"开盘价"
        dst_sheet['E2'] = u"换手率"
        dst_sheet['F1'] = u"卖出"
        dst_sheet.merge_cells("F1:H1")
        dst_sheet['F2'] = u"日期"
        dst_sheet['G2'] = u"开盘价"
        dst_sheet['H2'] = u"换手率"
        dst_sheet['I2'] = u"涨幅"
        dst_sheet['J2'] = u"每次交易平均换手率"
        dst_sheet['K2'] = u"总涨幅"

        capacity = 0
        capacity = len(data_in)
        start_row = 2
        interval = 0
        gain_totl = 0
        # print(data_in)
            
        for i in range(0, len(data_in)):
            # print(data_in[i])
            if not dst_sheet[u"A{}".format(start_row + 1 + interval)].value:
                start = start_row + 1 + interval
                end = start_row + 1 + interval + len(data_in)-1
                dst_sheet[u"A{}".format(start_row + 1 + interval)] = data_in[i][1].split("'")[1]
                dst_sheet.merge_cells("A{}:A{}".format(start, end))
                dst_sheet[u"B{}".format(start_row + 1 + interval)] = data_in[i][2]
                dst_sheet.merge_cells("B{}:B{}".format(start, end))
                dst_sheet.merge_cells("K{}:K{}".format(start, end))
            dst_sheet["C{}".format(start+i)] = data_in[i][0].strftime("%Y-%m-%d")
            dst_sheet["D{}".format(start+i)] = np.float64(data_in[i][3]).item()
            dst_sheet["E{}".format(start+i)] = np.float64(data_in[i][4]).item()
            if len(data_in) == len(data_out):
                dst_sheet["F{}".format(start+i)] = data_out[i][0].strftime("%Y-%m-%d")
                dst_sheet["G{}".format(start+i)] = np.float64(data_out[i][3]).item()
                dst_sheet["H{}".format(start+i)] = np.float64(data_out[i][4]).item()
                gain = np.float64(data_out[i][3]).item() - np.float64(data_in[i][3]).item()
                gain = gain / np.float64(data_in[i][3]).item()
                dst_sheet["I{}".format(start+i)] = gain
                # rate = (np.float64(data_in[i][4]).item() + np.float64(data_out[i][4]).item())/2
                rate_totl = 0
                # print("买入：{0}\n卖出:{1},天数：{2}".format( data_in[i], data_out[i], len(round_lst[i])) )

                for j in range(0, len(round_lst[i])):
                    # print("买入到卖出每天的换手率：", round_lst[i][j])
                    rate_totl += np.float64(round_lst[i][j][4]).item()
                # print("总换手率：", rate_totl)

                dst_sheet["J{}".format(start+i)] = rate_totl/len(round_lst[i])
                gain_totl += gain
            elif len(data_in) > len(data_out):
                if i == 0:
                    dst_sheet["F{}".format(start+i )] = None
                    dst_sheet["G{}".format(start+i)] = None
                    dst_sheet["H{}".format(start+i)] = None
                    dst_sheet["I{}".format(start+i)] = None
                    dst_sheet["J{}".format(start+i)] = None
                else:
                    dst_sheet["F{}".format(start+i )] = data_out[i-1][0].strftime("%Y-%m-%d")
                    dst_sheet["G{}".format(start+i )] = np.float64(data_out[i-1][3]).item()
                    dst_sheet["H{}".format(start+i)] = np.float64(data_out[i-1][4]).item()
                    gain = np.float64(data_out[i-1][3]).item() - np.float64(data_in[i][3]).item()
                    gain = gain / np.float64(data_in[i][3]).item()
                    dst_sheet["I{}".format(start+i)] = gain
                    # rate = (np.float64(data_in[i][4]).item() + np.float64(data_out[i-1][4]).item())/2
                    rate_totl = 0
                    # print("买入：{0}\n卖出:{1},天数：{2}".format( data_in[i], data_out[i-1], len(round_lst[i-1])) )

                    for j in range(0, len(round_lst[i-1])):
                        # print("买入到卖出每天的换手率：", round_lst[i-1][j])
                        rate_totl += np.float64(round_lst[i-1][j][4]).item()
                    # print("总换手率：", rate_totl)

                    dst_sheet["J{}".format(start+i)] = rate_totl/len(round_lst[i-1])
                    gain_totl += gain
        if not dst_sheet[u"K{}".format(start_row + 1 + interval)].value:
            dst_sheet[u"K{}".format(start_row + 1 + interval)] = gain_totl

        font = Font(u"宋体", size=12)
        for col in range(0, len(col_titles)):
            dst_sheet.column_dimensions["{}".format(col_titles[col])].width = col_widthes[col]
            for row in range(1, dst_sheet.max_row + 1):
                dst_sheet[u"{}{}".format(col_titles[col], row)].font = font
                dst_sheet[u"{}{}".format(col_titles[col], row)].alignment = Alignment(horizontal='center', vertical='center')

        filename = '001.xlsx'
        dst_wb.save(filename)
        
def main():
    parser = argparse.ArgumentParser(description='读取CSV文件')
    parser.add_argument('-n', dest='name',  help='文件名', default='default')
    args = parser.parse_args()
    f = ReadAllCSV()
    if args.name == "all":
        f.getTotalIncreaseCSV()
    else:
        f.getSigleCSV(args.name)

if __name__ == "__main__":
    main()