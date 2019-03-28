# -*- coding: utf-8 -*-
import csv
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, colors, Border, Side, PatternFill, Alignment
import time
import numpy as np
def get_file_path():
    # 获取脚本路径
    path = sys.path[0]
    # 判断为脚本文件还是py2exe编译后的文件，如果是脚本文件，则返回的是脚本的目录，如果是py2exe编译后的文件，则返回的是编译后的文件路径
    if os.path.isdir(path):
        return path
    elif os.path.isfile(path):
        return os.path.dirname(path)

def batch_read(fillename, batch_size=5):
    with open(file=fillename, mode='rb') as fd:
        batch_output = []
        for (n, row) in enumerate(csv.reader(fd, dialect='excel')):
                # 当填满一个batch时,yield这个batch
                if n > 0 and n % batch_size == 0:
                    # 填满batch后才yield
                    yield(np.array(batch_output))
                    # 重置batch
                    batch_output = list()
                batch_output.append(row)
            # 在最后yield剩余的数据
            # 只有在没有填满一个batch的情况下才会执行此条语句
        yield(np.array(batch_output))

RUSHOUSHIJIAN = 7
LIANXU = 4
HUANSHOULV = -1
CHUSHOUSHIJIAN = 7
def test():
    path = os.path.join(get_file_path(), "resource\Data")
    files = os.listdir(path)
    file_name = os.path.join(get_file_path(), "resource\Data\{}".format("data600086.csv"))
    
    target_in = []
    ishold = False
    target_out = []
    with open(file=file_name, encoding='ANSI') as fd:
        df = pd.read_csv(fd, sep=',', encoding='utf-8', usecols=['日期', '股票代码', '名称', '开盘价', '换手率'])
        df['日期'] = pd.to_datetime(df['日期'], format="%Y-%m-%d")
        df.sort_values(by='日期', inplace=True)
        # print(df)

        company = {}
        in_seven_lst = []
        four_lst = []
        out_seven_lst = []
        for indexs in df.index: 
            data_everyday = df.loc[indexs].values
            #个别无数据的跳过
            if np.float64(data_everyday[3]).item() == 0:
                continue
            isrise = False
            in_seven_lst.insert(0, data_everyday)
            out_seven_lst.insert(0, data_everyday)

            if not ishold:
                if len(in_seven_lst) == RUSHOUSHIJIAN + 1:
                    objective = in_seven_lst[0]   #用来比较的对象
                    average = 0
                    total = 0
                    for i in range(1, len(in_seven_lst)):
                        total += float(in_seven_lst[i][HUANSHOULV])
                    average = total/RUSHOUSHIJIAN
                    
                    #判断换手率大于7天平均值
                    if float(objective[HUANSHOULV]) > average and average != 0:
                        four_lst.insert(0, objective)
                        isrise = True
                    #买入 确保买入的时间是连续4天上升的第二天
                    if len(four_lst) == LIANXU + 1:
                        target_in.insert(0, objective)
                        ishold = True
                    if not isrise or ishold:
                        four_lst.clear()
            else:
                if len(out_seven_lst) == CHUSHOUSHIJIAN+1:
                    out_objective = out_seven_lst[0]
                    out_average = 0
                    out_totle = 0
                    for j in range(1, len(out_seven_lst)):
                        out_totle += float(out_seven_lst[j][HUANSHOULV])
                    out_average = out_totle/CHUSHOUSHIJIAN

                    if float(out_objective[HUANSHOULV]) < out_average and out_average != 0:
                        target_out.insert(0, out_objective)
                        ishold = False
            if len(in_seven_lst) == RUSHOUSHIJIAN + 1:
                del in_seven_lst[-1]  #删除最后一个元素
            if len(out_seven_lst) == CHUSHOUSHIJIAN+1:
                del out_seven_lst[-1]
    writecsv(target_in, target_out)

def writecsv(data_in, data_out):
    col_titles = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    col_widthes = [8, 12, 12, 12, 12, 12, 12, 12,12, 12, 12]

    dst_wb = openpyxl.Workbook()
    dst_sheet = dst_wb.active
    dst_sheet.title = u"股票统计"
    dst_sheet['A2'] = u"股票代码"
    dst_sheet['B2'] = u"名称"
    dst_sheet['C1'] = u"买入"
    dst_sheet.merge_cells("C1:E1")
    dst_sheet['C2'] = u"日期"
    dst_sheet['D2'] = u"开盘价"
    dst_sheet['E2'] = u"换手率"
    dst_sheet['F1'] = u"卖出"
    dst_sheet.merge_cells("F1:H1")
    dst_sheet['F2'] = u"日期"
    dst_sheet['G2'] = u"开盘价"
    dst_sheet['H2'] = u"换手率"
    dst_sheet['I2'] = u"涨幅"
    dst_sheet['J2'] = u"每次交易平均换手率"
    dst_sheet['K2'] = u"总涨幅"

    capacity = 0
    capacity = len(data_in)
    start_row = 2
    interval = 0
    gain_totl = 0
    # print(data_in)
    for i in range(0, len(data_in)):
        # print(data_in[i])
        if not dst_sheet[u"A{}".format(start_row + 1 + interval)].value:
            start = start_row + 1 + interval
            end = start_row + 1 + interval + len(data_in)-1
            dst_sheet[u"A{}".format(start_row + 1 + interval)] = data_in[i][1].split("'")[1]
            dst_sheet.merge_cells("A{}:A{}".format(start, end))
            dst_sheet[u"B{}".format(start_row + 1 + interval)] = data_in[i][2]
            dst_sheet.merge_cells("B{}:B{}".format(start, end))
            dst_sheet.merge_cells("K{}:K{}".format(start, end))
        dst_sheet["C{}".format(start+i)] = data_in[i][0].strftime("%Y-%m-%d")
        dst_sheet["D{}".format(start+i)] = np.float64(data_in[i][3]).item()
        dst_sheet["E{}".format(start+i)] = np.float64(data_in[i][4]).item()
        if i < len(data_out):
            dst_sheet["F{}".format(start+i)] = data_out[i][0].strftime("%Y-%m-%d")
            dst_sheet["G{}".format(start+i)] = np.float64(data_out[i][3]).item()
            dst_sheet["H{}".format(start+i)] = np.float64(data_out[i][4]).item()
            gain = np.float64(data_out[i][3]).item() - np.float64(data_in[i][3]).item()
            dst_sheet["I{}".format(start+i)] = gain
            rate = (np.float64(data_in[i][4]).item() + np.float64(data_out[i][4]).item())/2
            dst_sheet["J{}".format(start+i)] = rate
            gain_totl += gain
        if not dst_sheet[u"K{}".format(start_row + 1 + interval)].value:
            dst_sheet[u"K{}".format(start_row + 1 + interval)] = gain_totl

    font = Font(u"宋体", size=12)
    for col in range(0, len(col_titles)):
        dst_sheet.column_dimensions["{}".format(col_titles[col])].width = col_widthes[col]
        for row in range(1, dst_sheet.max_row + 1):
            dst_sheet[u"{}{}".format(col_titles[col], row)].font = font
            dst_sheet[u"{}{}".format(col_titles[col], row)].alignment = Alignment(horizontal='center', vertical='center')

    filename = '001.xlsx'
    dst_wb.save(filename)

def test1():
    path = os.path.join(get_file_path(), "resource\Data")
    files = os.listdir(path)
    file_name = os.path.join(get_file_path(), "resource\Data\{}".format("data600010.csv"))
    
    target_in = []
    ishold = False
    target_out = []
    with open(file=file_name, encoding='ANSI') as fd:
        # data = pd.read_csv(fd, sep=',', encoding='utf-8', usecols=['日期', '股票代码', '名称', '涨跌幅', '换手率'])
        df = pd.read_csv(fd, sep=',', encoding='utf-8', usecols=['日期', '股票代码', '名称', '涨跌幅', '换手率'])
        df['日期'] = pd.to_datetime(df['日期'])
        df.sort_values(by='日期', inplace=True)
        # print(df)
        for indexs in df.index: 
            data_everyday = df.loc[indexs].values
            print(data_everyday)
            target_in.insert(0, data_everyday)            
    print(target_in[0][1])
    writecsv(target_in,target_out)

if __name__ == "__main__":
    test()